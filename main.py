import requests
import customtkinter as ctk
import tkinter as tk
import json
import os
import traceback
import threading
import time
from decimal import Decimal, ROUND_DOWN
import ccxt
import openpyxl
from datetime import datetime
from tkinter import messagebox

# Configuración
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

CONFIG_FILE = "data/config.json"
HISTORY_DIR = "data"
HISTORY_FILE_TEMPLATE = "trading_history_{mes}.xlsx"

EXCHANGES = {
    "Bybit": "bybit",
    "Binance": "binance",
    "Coinbase Exchange": "coinbaseexchange",
    "OKX": "okx",
    "Bitget": "bitget",
    "MEXC": "mexc",
    "KuCoin": "kucoin",
    "BingX": "bingx",
}

TOP_PAIRS = [
    "BTC/USDT",
    "ETH/USDT",
    "LTC/USDT",
    "DOT/USDT",
    "USDC/USDT",
    "SOL/USDT",
    "XRP/USDT",
    "RUNE/USDT",
    "TOMI/USDT",
    "DAI/USDT",
    "ROSE/USDT",
    "DOGE/USDT",
    "ADA/USDT",
    "TON/USDT",
    "AVAX/USDT",
    "XAUT/USDT",
    "ARB/USDT",
    "USDE/USDT",
    "FIL/USDT",
    "BNB/USDT",
    "LINK/USDT",
    "ALGO/USDT",
    "BCH/USDT",
    "EOS/USDT",
    "ETC/USDT",
    "UNI/USDT",
    "ATOM/USDT",
    "MANA/USDT",
    "SAND/USDT",
    "THETA/USDT",
    "XTZ/USDT",
    "BTC/EUR",
    "USDT/EUR",
    "ETH/EUR",
    "USDC/EUR",
    "BTC/BRL",
    "ETH/BRL",
]

trading_history = []


def save_config(data):
    os.makedirs(HISTORY_DIR, exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump(data, f, indent=4)


def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            return json.load(f)
    return {}


def get_current_history_file():
    fecha_mes = datetime.now().strftime("%m-%Y")
    return os.path.join(HISTORY_DIR, HISTORY_FILE_TEMPLATE.format(mes=fecha_mes))


def save_trading_history():
    os.makedirs(HISTORY_DIR, exist_ok=True)
    history_file = get_current_history_file()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"
    ws.append(
        [
            "#",
            "Exchange",
            "Par",
            "Fecha y Hora",
            "Precio Entrada",
            "Precio Salida",
            "Monto Total",
            "Fee Compra",
            "Fee Venta",
            "PnL",
        ]
    )

    total_pnl = 0
    for idx, t in enumerate(trading_history, start=1):
        ws.append(
            [
                idx,
                t["exchange"],
                t["symbol"],
                t["datetime"],
                t["entry_price"],
                t["exit_price"],
                t["amount"],
                t["fee_buy"],
                t["fee_sell"],
                t["pnl"],
            ]
        )
        total_pnl += t["pnl"]

    ws.append(["", "", "", "", "", "", "", "",
              "Total PnL", round(total_pnl, 6)])
    wb.save(history_file)


def load_trading_history():
    history_file = get_current_history_file()
    if os.path.exists(history_file):
        wb = openpyxl.load_workbook(history_file)
        ws = wb.active
        history = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and isinstance(row[9], (int, float)):
                history.append(
                    {
                        "exchange": row[1],
                        "symbol": row[2],
                        "datetime": row[3],
                        "entry_price": float(row[4]),
                        "exit_price": float(row[5]),
                        "amount": float(row[6]),
                        "fee_buy": float(row[7]),
                        "fee_sell": float(row[8]),
                        "pnl": float(row[9]),
                    }
                )
        return history
    return []


def ajustar_qty(qty, step_size):
    step = Decimal(str(step_size))
    return float((Decimal(str(qty)).quantize(step, rounding=ROUND_DOWN)))


# -------------------------------------------------------
# Clase BotFrame (con botones abajo y etiquetas arriba una sola vez)
# -------------------------------------------------------


class BotFrame(ctk.CTkFrame):
    # Controla que las etiquetas de botones se muestren solo una vez
    etiquetas_mostradas = False

    def __init__(
        self,
        master,
        symbol,
        monto,
        tp_pct,
        sep_pct,
        os_num,
        exchange,
        remove_callback,
        numero=None,
    ):
        super().__init__(master, fg_color="transparent")

        self.symbol = symbol
        self.monto = monto
        self.tp_pct = tp_pct
        self.sep_pct = sep_pct
        self.os_num = os_num
        self.exchange = exchange
        self.exchange_id = exchange.id
        self.running = True
        self.remove_callback = remove_callback
        self.app = self._get_app()

        self.contador_ciclos = 0
        self.total_trades = 0

        self.reiniciar_var = tk.BooleanVar(value=False)
        self.reiniciar_os_var = tk.BooleanVar(value=False)

        self.tp_orders = []
        self.entry_time = None
        self.vistos_buys = set()
        self.order_base_id = None

        self.vnc_total = 0.0
        self.vnc_total_cost = 0.0

        self.dca_sell_pct_var = tk.StringVar(value="0.5%")
        self.sl_pct_var = tk.StringVar(value="")

        # Frame principal de cada bot
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.pack(fill="x", padx=10, pady=(3, 0))

        # Línea superior: texto a la izquierda, botones a la derecha
        self.top_row_frame = ctk.CTkFrame(
            self.main_frame, fg_color="transparent")
        self.top_row_frame.pack(fill="x", padx=10, pady=(3, 0))

        # Texto informativo del bot (izquierda)
        self.label_info = ctk.CTkLabel(
            self.top_row_frame, text="", font=("Arial", 12, "bold"), text_color="white"
        )
        self.label_info.pack(side="left", pady=5)

        # Controles del bot (derecha)
        self.bot_controls_frame = ctk.CTkFrame(
            self.top_row_frame, fg_color="transparent"
        )
        self.bot_controls_frame.pack(side="right", padx=(0, 0))

        # Caja de texto SL (reemplaza el menú)
        self.sl_entry = ctk.CTkEntry(
            self.bot_controls_frame,
            textvariable=self.sl_pct_var,
            width=60,
            font=("Arial", 10),
            justify="center",
        )
        self.sl_entry.pack(side="left", padx=(1, 1))

        # Caja de texto TP+ (nuevo campo)
        self.tp_plus_var = tk.StringVar(value="")
        self.tp_plus_entry = ctk.CTkEntry(
            self.bot_controls_frame,
            textvariable=self.tp_plus_var,
            width=60,
            font=("Arial", 10),
            justify="center",
        )
        self.tp_plus_entry.pack(side="left", padx=(1, 1))

        # Menú TP VNC
        self.dca_menu = ctk.CTkOptionMenu(
            self.bot_controls_frame,
            values=["0.1%", "0.2%", "0.5%", "1%",
                    "2%", "3%", "5%", "10%", "20%"],
            variable=self.dca_sell_pct_var,
            width=60,
            font=("Arial", 10),
            anchor="center",
        )
        self.dca_menu.pack(side="left", padx=(1, 1))

        # Checkbox R
        self.reiniciar_checkbox = ctk.CTkCheckBox(
            self.bot_controls_frame, text="", variable=self.reiniciar_var, width=0
        )
        self.reiniciar_checkbox.pack(side="left", padx=(1, 1))

        # Checkbox R(OS)
        self.reiniciar_os_checkbox = ctk.CTkCheckBox(
            self.bot_controls_frame, text="", variable=self.reiniciar_os_var, width=0
        )
        self.reiniciar_os_checkbox.pack(side="left", padx=(1, 1))

        # Botón Detener
        self.stop_button = ctk.CTkButton(
            self.bot_controls_frame,
            text="🛑",
            width=40,
            height=25,
            command=self.stop_bot,
        )
        self.stop_button.pack(side="left", padx=(1, 1))

        self.separator = ctk.CTkLabel(
            self, text="─" * 300, text_color="#555555")
        self.separator.pack(fill="x", padx=0, pady=(2, 0))

        self.update_labels()
        threading.Thread(target=self.run_bot, daemon=True).start()
        threading.Thread(
            target=self.actualizar_precio_tiempo_real, daemon=True).start()

    def _get_app(self):
        parent = self.master
        while parent and not isinstance(parent, BotDCAApp):
            parent = getattr(parent, "master", None)
        return parent

    def update_labels(self, precio_actual=None):
        dca = self.vnc_total_cost / self.vnc_total if self.vnc_total > 0 else 0
        symbol_base = self.symbol.split("/")[0]
        precio_actual_texto = (
            f" | (PA): {format(precio_actual, '.4f')}" if precio_actual else ""
        )
        info_text = (
            f"{self.exchange.id.upper()} | {self.symbol} | TP: {self.tp_pct}% | MO: {self.monto} USDT | "
            f"OS: {self.os_num} | OC: {self.contador_ciclos} | TTR: {self.total_trades} | "
            f"VNC: {format(self.vnc_total, '.8f')} {symbol_base} | DCA: {format(dca, '.4f')}{precio_actual_texto}"
        )
        self.label_info.configure(text=info_text, text_color="white")

    def actualizar_precio_tiempo_real(self):
        while self.running:
            try:
                precio_actual = self.exchange.fetch_ticker(self.symbol)["last"]
                self.update_labels(precio_actual)

                if self.vnc_total > 0:
                    dca = self.vnc_total_cost / self.vnc_total
                    selected_pct = (
                        float(self.dca_sell_pct_var.get().replace("%", "")) / 100
                    )
                    trigger_price = dca * (1 + selected_pct)

                    if precio_actual >= trigger_price:
                        print(
                            f"[AUTOSELL] Vendiendo {self.vnc_total:.8f} a mercado a {precio_actual:.4f}")
                    try:
                        symbol_base = self.symbol.split("/")[0]
                        balance = self.exchange.fetch_balance()
                        qty = balance.get(symbol_base, {}).get("free", 0)
                        qty = ajustar_qty(
                            qty,
                            self.exchange.markets[self.symbol]["precision"]["amount"],
                        )

                        if qty > 0:
                            self.exchange.create_order(
                                symbol=self.symbol,
                                type="market",
                                side="sell",
                                amount=qty,
                            )

                            fee_buy = 0.0
                            fee_sell = 0.0
                            pnl = (precio_actual - dca) * \
                                qty - fee_buy - fee_sell
                            fecha_hora = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

                            trading_history.append(
                                {
                                    "exchange": self.exchange.id.upper(),
                                    "symbol": self.symbol,
                                    "datetime": fecha_hora,
                                    "entry_price": round(dca, 4),
                                    "exit_price": round(precio_actual, 4),
                                    "amount": round(qty, 6),
                                    "fee_buy": round(fee_buy, 6),
                                    "fee_sell": round(fee_sell, 6),
                                    "pnl": round(pnl, 6),
                                }
                            )

                            save_trading_history()
                            self.app.session_trades += 1
                            self.app.update_dashboard()

                            self.vnc_total = 0.0
                            self.vnc_total_cost = 0.0
                        else:
                            print(
                                "[AUTOSELL] No hay balance libre disponible para vender.")
                    except Exception as e:
                        print(f"[ERROR AUTOSELL] {e}")

                        # Verificación de Stop Loss con valor fijo en caja de texto

                sl_value = self.sl_pct_var.get().strip()
                if sl_value:
                    try:
                        sl_price = float(sl_value)

                        if precio_actual <= sl_price:
                            print(
                                f"[STOP LOSS] Precio actual {precio_actual:.4f} <= SL {sl_price:.4f}")

                        # 1. Deseleccionar los checkboxes de reinicio
                        self.reiniciar_var.set(False)
                        self.reiniciar_checkbox.deselect()
                        self.reiniciar_os_var.set(False)
                        self.reiniciar_os_checkbox.deselect()

                        # 2. Detener el bot como si se presionara el botón 🛑
                        self.stop_bot()

                        # 3. Esperar a que todas las órdenes estén cerradas (máx 8 segundos)
                        timeout = time.time() + 8
                        while time.time() < timeout:
                            try:
                                open_orders = self.exchange.fetch_open_orders(
                                    self.symbol)
                                if not open_orders:
                                    break
                                print("[SL] Esperando cierre de órdenes...")
                                time.sleep(2)
                            except Exception as e:
                                print(
                                    f"[SL] Error al verificar órdenes abiertas: {e}")
                                break

                        # 4. Vender todo el balance disponible de esa cripto a precio de mercado
                        try:
                            symbol_base = self.symbol.split("/")[0]
                            balance = self.exchange.fetch_balance()
                            qty = balance.get(symbol_base, {}).get("free", 0)
                            qty = ajustar_qty(
                                qty,
                                self.exchange.markets[self.symbol]["precision"]["amount"],
                            )

                            if qty > 0:
                                self.exchange.create_order(
                                    symbol=self.symbol,
                                    type="market",
                                    side="sell",
                                    amount=qty,
                                )

                                entry_price = self.vnc_total_cost / self.vnc_total if self.vnc_total > 0 else 0
                                fee_buy = 0.0
                                fee_sell = 0.0
                                pnl = (precio_actual - entry_price) * \
                                    qty - fee_buy - fee_sell
                                fecha_hora = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

                                trading_history.append(
                                    {
                                        "exchange": self.exchange.id.upper(),
                                        "symbol": self.symbol,
                                        "datetime": fecha_hora,
                                        "entry_price": round(entry_price, 4),
                                        "exit_price": round(precio_actual, 4),
                                        "amount": round(qty, 6),
                                        "fee_buy": round(fee_buy, 6),
                                        "fee_sell": round(fee_sell, 6),
                                        "pnl": round(pnl, 6),
                                    }
                                )

                                save_trading_history()
                                self.app.session_trades += 1
                                self.app.update_dashboard()
                                print(
                                    f"[SL] Vendido todo el balance disponible: {qty} a mercado")
                            else:
                                print(
                                    "[SL] No hay balance disponible para vender.")
                        except Exception as e:
                            print(f"[ERROR SL] {e}")

                        return  # Finaliza el hilo del bot tras SL

                    except Exception as e:
                        print(f"[ERROR SL] {e}")

                # Verificación de TP+ (nuevo)
                tp_plus_value = self.tp_plus_var.get().strip()
                if tp_plus_value:
                    try:
                        tp_plus_price = float(tp_plus_value)
                        if precio_actual >= tp_plus_price:
                            print(
                                f"[TP+] Precio actual {precio_actual:.4f} ≥ TP+ {tp_plus_price:.4f}")

                            # 1. Deseleccionar checkboxes
                            self.reiniciar_var.set(False)
                            self.reiniciar_checkbox.deselect()
                            self.reiniciar_os_var.set(False)
                            self.reiniciar_os_checkbox.deselect()

                            # 2. Detener bot
                            self.stop_bot()

                            # 3. Esperar 10 segundos
                            time.sleep(10)

                            # 4. Vender todo el balance disponible a mercado
                            symbol_base = self.symbol.split("/")[0]
                            balance = self.exchange.fetch_balance()
                            qty = balance.get(symbol_base, {}).get("free", 0)
                            qty = ajustar_qty(
                                qty,
                                self.exchange.markets[self.symbol]["precision"]["amount"],
                            )

                            if qty > 0:
                                self.exchange.create_order(
                                    symbol=self.symbol,
                                    type="market",
                                    side="sell",
                                    amount=qty,
                                )

                                entry_price = self.vnc_total_cost / self.vnc_total if self.vnc_total > 0 else 0
                                fee_buy = 0.0
                                fee_sell = 0.0
                                pnl = (precio_actual - entry_price) * \
                                    qty - fee_buy - fee_sell
                                fecha_hora = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

                                trading_history.append(
                                    {
                                        "exchange": self.exchange.id.upper(),
                                        "symbol": self.symbol,
                                        "datetime": fecha_hora,
                                        "entry_price": round(entry_price, 4),
                                        "exit_price": round(precio_actual, 4),
                                        "amount": round(qty, 6),
                                        "fee_buy": round(fee_buy, 6),
                                        "fee_sell": round(fee_sell, 6),
                                        "pnl": round(pnl, 6),
                                    }
                                )

                                save_trading_history()
                                self.app.session_trades += 1
                                self.app.update_dashboard()

                                self.vnc_total = 0.0
                                self.vnc_total_cost = 0.0
                                print(f"[TP+] Vendido {qty} a mercado")
                            else:
                                print("[TP+] No hay balance para vender")
                    except Exception as e:
                        print(f"[ERROR TP+] {e}")

            except Exception as e:
                print(f"[Error] al obtener precio en tiempo real: {e}")
            time.sleep(2)

    def run_bot(self):
        try:
            self.exchange.load_markets()
            market = self.exchange.markets[self.symbol]
            last_price = self.exchange.fetch_ticker(self.symbol)["last"]
            step_size = market["precision"]["amount"]
            price_precision = market["precision"]["price"]
            tick_size = 1 / (10**price_precision)
            min_cost = market["limits"]["cost"]["min"]

            qty = ajustar_qty(self.monto / last_price, step_size)
            order_value = qty * last_price

            if order_value < min_cost:
                self.label_info.configure(
                    text=f"❌ Monto insuficiente en {self.symbol}", text_color="red"
                )
                return

            order = self.exchange.create_order(
                symbol=self.symbol, type="market", side="buy", amount=qty
            )
            self.order_base_id = order["id"]
            self.entry_time = int(time.time() * 1000)
            self.vistos_buys.add(order["id"])

            self.colocar_tp(qty, last_price)

            for i in range(1, self.os_num + 1):
                precio_os = last_price * (1 - (self.sep_pct / 100) * i)
                precio_os = float(
                    Decimal(precio_os).quantize(
                        Decimal(str(tick_size)), rounding=ROUND_DOWN
                    )
                )
                qty_os = ajustar_qty(self.monto / precio_os, step_size)
                self.exchange.create_order(
                    symbol=self.symbol,
                    type="limit",
                    side="buy",
                    amount=qty_os,
                    price=precio_os,
                )

            self.monitor_ordenes()
        except Exception:
            traceback.print_exc()

    def colocar_tp(self, qty, base_price):
        try:
            market = self.exchange.markets[self.symbol]
            price_precision = market["precision"]["price"]
            tick_size = 1 / (10**price_precision)

            tp_price = float(
                Decimal(base_price * (1 + self.tp_pct / 100)).quantize(
                    Decimal(str(tick_size)), rounding=ROUND_DOWN
                )
            )
            qty = ajustar_qty(qty, market["precision"]["amount"])

            tp_order = self.exchange.create_order(
                symbol=self.symbol,
                type="limit",
                side="sell",
                amount=qty,
                price=tp_price,
            )
            self.tp_orders.append(tp_order["id"])
        except Exception:
            traceback.print_exc()

    def monitor_ordenes(self):
        try:
            while self.running:
                time.sleep(2)
                try:
                    open_orders = self.exchange.fetch_open_orders(self.symbol)
                except Exception as e:
                    print(f"[Error] fetch_open_orders: {e}")
                    continue

                try:
                    sell_orders = [
                        o for o in open_orders if o["side"] == "sell"]
                    buy_orders = [o for o in open_orders if o["side"] == "buy"]

                    if self.reiniciar_var.get() and not sell_orders:
                        for order in buy_orders:
                            try:
                                self.exchange.cancel_order(
                                    order["id"], self.symbol)
                            except Exception:
                                pass
                        self.running = False
                        self.reiniciar_bot(origen="TP")
                        return

                    if self.reiniciar_os_var.get() and not buy_orders:
                        for order in sell_orders:
                            try:
                                self.exchange.cancel_order(
                                    order["id"], self.symbol)
                            except Exception:
                                pass
                        self.running = False
                        self.reiniciar_bot(origen="OS")
                        return

                    try:
                        filled_trades = self.exchange.fetch_my_trades(
                            self.symbol)
                    except Exception as e:
                        print(f"[Error] fetch_my_trades: {e}")
                        continue

                    new_trades = [
                        t for t in filled_trades if t["timestamp"] >= self.entry_time
                    ]

                    for trade in new_trades:
                        try:
                            if (
                                trade["side"] == "buy"
                                and trade["order"] not in self.vistos_buys
                            ):
                                self.vistos_buys.add(trade["order"])
                                qty = ajustar_qty(
                                    trade["amount"],
                                    self.exchange.markets[self.symbol]["precision"][
                                        "amount"
                                    ],
                                )
                                base_price = trade["price"]
                                self.colocar_tp(qty, base_price)

                            if (
                                trade["side"] == "sell"
                                and trade["order"] in self.tp_orders
                            ):
                                self.tp_orders.remove(trade["order"])
                                self.total_trades += 1
                                self.app.session_trades += 1
                                self.update_labels()
                                self.registrar_trade(trade)
                                self.colocar_nueva_os(trade["price"])
                        except Exception as e:
                            print(f"[Error] procesando trade: {e}")
                except Exception as e:
                    print(f"[Error] en monitor_ordenes: {e}")
        except Exception:
            traceback.print_exc()

    def registrar_trade(self, trade):
        try:
            fecha_hora = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            symbol = trade["symbol"]
            exit_price = trade["price"]
            amount = trade["amount"]

            # Usamos el DCA acumulado como entry_price si hay VNC
            if self.vnc_total > 0:
                entry_price = self.vnc_total_cost / self.vnc_total
            else:
                entry_price = trade.get("cost", 0) / trade.get("amount", 1)

            # Estimamos fee de compra si no hay datos
            fee_buy = 0.001 * entry_price * amount  # estimado 0.1%

            # Fee venta desde el trade o estimado si no está
            fee_sell = trade.get("fee", {}).get("cost", 0)
            if not fee_sell:
                fee_sell = 0.001 * exit_price * amount  # estimado 0.1%

            # Calculamos PnL real
            pnl = (exit_price - entry_price) * amount - fee_buy - fee_sell

            trading_history.append(
                {
                    "exchange": self.exchange.id.upper(),
                    "symbol": symbol,
                    "datetime": fecha_hora,
                    "entry_price": round(entry_price, 4),
                    "exit_price": round(exit_price, 4),
                    "amount": round(amount, 6),
                    "fee_buy": round(fee_buy, 6),
                    "fee_sell": round(fee_sell, 6),
                    "pnl": round(pnl, 6),
                }
            )

            save_trading_history()
            self.app.update_dashboard()
        except Exception:
            traceback.print_exc()

    def colocar_nueva_os(self, last_sell_price):
        try:
            market = self.exchange.markets[self.symbol]
            price_precision = market["precision"]["price"]
            tick_size = 1 / (10**price_precision)

            new_buy_price = last_sell_price * (1 - (self.sep_pct / 100))
            new_buy_price = float(
                Decimal(new_buy_price).quantize(
                    Decimal(str(tick_size)), rounding=ROUND_DOWN
                )
            )
            qty = ajustar_qty(self.monto / new_buy_price,
                              market["precision"]["amount"])

            self.exchange.create_order(
                symbol=self.symbol,
                type="limit",
                side="buy",
                amount=qty,
                price=new_buy_price,
            )
        except Exception:
            traceback.print_exc()

    def reiniciar_bot(self, origen="TP"):
        try:
            if origen == "OS":
                filled_trades = self.exchange.fetch_my_trades(self.symbol)
                nuevos_buys = [
                    t
                    for t in filled_trades
                    if t["timestamp"] >= self.entry_time and t["side"] == "buy"
                ]

                symbol_base = self.symbol.split("/")[0]
                total_qty = 0.0
                total_cost = 0.0

                for buy in nuevos_buys:
                    qty = float(buy["amount"])
                    price = float(buy["price"])
                    total_qty += qty
                    total_cost += qty * price

                if total_qty > 0:
                    self.vnc_total += total_qty
                    self.vnc_total_cost += total_cost
                    print(
                        f"[REINICIO-OS] +{total_qty:.8f} {symbol_base} acumulados | Costo: {total_cost:.4f} | Nuevo VNC: {self.vnc_total:.8f}"
                    )
        except Exception as e:
            print(f"[Error] al calcular VNC/DCA: {e}")

            self.contador_ciclos += 1

            # Aquí empieza el nuevo bloque:
        self.entry_time = int(time.time() * 1000)
        self.vistos_buys = set()
        self.tp_orders.clear()
        self.running = True

        threading.Thread(target=self.run_bot, daemon=True).start()

        self.reiniciar_checkbox.configure(fg_color="green")
        self.after(1000, lambda: self.reiniciar_checkbox.configure(
            fg_color="#1f6aa5"))

    def stop_bot(self):
        self.dca_sell_pct_var.set("0.5%")
        self.sl_pct_var.set("0%")
        threading.Thread(target=self._stop_bot_thread, daemon=True).start()

    def _stop_bot_thread(self):
        try:
            self.running = False
            try:
                open_orders = self.exchange.fetch_open_orders(self.symbol)
                for order in open_orders:
                    try:
                        self.exchange.cancel_order(order["id"], self.symbol)
                    except Exception:
                        pass
            except Exception as e:
                print(f"[Error] al cancelar órdenes al detener: {e}")

            self.label_info.configure(
                text="🛑 Bot detenido correctamente", text_color="orange"
            )
            time.sleep(1)
            self.after(100, self.destroy)
            self.remove_callback(self)

          # ✅ Detener también el bot en el servidor

            try:
                requests.post("https://bot-server-cg3g.onrender.com/api/stop_bot",
                              json={"id": f"{self.symbol}-{self.exchange_id}"})
                print(
                    f"🛑 Bot detenido en servidor: {self.symbol}-{self.exchange_id}")
            except Exception as e:
                print(f"[SERVER STOP ERROR] {e}")

            self.app.update_dashboard()

        except Exception:
            traceback.print_exc()

    def show_tooltip(self, text):
        self.tooltip = ctk.CTkLabel(
            self, text=text, bg_color="black", text_color="white", font=("Arial", 10)
        )
        self.tooltip.place(
            x=self.winfo_pointerx() - self.winfo_rootx(),
            y=self.winfo_pointery() - self.winfo_rooty() + 20,
        )

    def hide_tooltip(self):
        if hasattr(self, "tooltip"):
            self.tooltip.destroy()


# -------------------------------------------------------
# Clase Principal: BotDCAApp
# -------------------------------------------------------


class BotDCAApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("BotDCA PRO Multi-Exchange")
        self.geometry("1100x700")
        self.config = load_config()
        self.exchange_instances = {}
        self.bots = []
        self.session_trades = 0

        global trading_history
        trading_history = load_trading_history()

        self.main_frame = ctk.CTkFrame(self)
        self.main_frame.pack(fill="both", expand=True)

        # Panel superior: conexión
        self.connection_frame = ctk.CTkFrame(self.main_frame)
        self.connection_frame.pack(pady=5, padx=10, fill="x")

        self.exchange_var = tk.StringVar()
        self.exchange_menu = ctk.CTkOptionMenu(
            self.connection_frame,
            values=list(EXCHANGES.keys()),
            variable=self.exchange_var,
            command=self.load_exchange_api,
        )
        self.exchange_menu.pack(side="left", padx=5)

        self.api_key_entry = ctk.CTkEntry(
            self.connection_frame, placeholder_text="API Key"
        )
        self.api_key_entry.pack(side="left", padx=5, fill="x", expand=True)

        self.api_secret_entry = ctk.CTkEntry(
            self.connection_frame, placeholder_text="API Secret", show="*"
        )
        self.api_secret_entry.pack(side="left", padx=5, fill="x", expand=True)

        self.connect_button = ctk.CTkButton(
            self.connection_frame, text="Conectar", command=self.toggle_connection
        )
        self.connect_button.pack(side="left", padx=5)

        self.status_circle = ctk.CTkLabel(
            self.connection_frame, text="⚪", text_color="gray"
        )
        self.status_circle.pack(side="left", padx=10)

        # Panel de configuración del bot
        self.controls_frame = ctk.CTkFrame(self.main_frame)
        self.controls_frame.pack(pady=5, padx=10, fill="x")

        self.pair_var = tk.StringVar()
        self.pair_dropdown = ctk.CTkComboBox(
            self.controls_frame, values=TOP_PAIRS, variable=self.pair_var
        )
        self.pair_dropdown.pack(side="left", padx=5)

        self.tp_entry = ctk.CTkEntry(
            self.controls_frame, placeholder_text="TP%")
        self.tp_entry.pack(side="left", padx=5)

        self.sep_entry = ctk.CTkEntry(
            self.controls_frame, placeholder_text="Separación%"
        )
        self.sep_entry.pack(side="left", padx=5)

        self.os_entry = ctk.CTkEntry(
            self.controls_frame, placeholder_text="OS")
        self.os_entry.pack(side="left", padx=5)

        self.monto_entry = ctk.CTkEntry(
            self.controls_frame, placeholder_text="Monto")
        self.monto_entry.pack(side="left", padx=5)

        self.start_button = ctk.CTkButton(
            self.main_frame,
            text="▶ Iniciar Bot",
            corner_radius=10,
            command=self.start_bot,
        )
        self.start_button.pack(pady=(10, 5))

        self.historial_button = ctk.CTkButton(
            self.main_frame, text="📈 Historial de trading", command=self.abrir_historial
        )
        self.historial_button.pack(pady=(0, 10))

        # Etiquetas de los controles del bot (R, R(OS), SL, etc.) al lado del botón de historial
        self.titulo_bot_controls = ctk.CTkFrame(
            self.main_frame, fg_color="transparent")
        self.titulo_bot_controls.pack(fill="x", padx=10, pady=(0, 5))

        ctk.CTkLabel(
            self.titulo_bot_controls, text="🛑", font=("Arial", 25, "bold")
        ).pack(side="right", padx=25)
        ctk.CTkLabel(
            self.titulo_bot_controls, text="R(OS)", font=("Arial", 13, "bold")
        ).pack(side="right", padx=3)
        ctk.CTkLabel(
            self.titulo_bot_controls, text="R-", font=("Arial", 13, "bold")
        ).pack(side="right", padx=5)
        ctk.CTkLabel(
            self.titulo_bot_controls, text="TP VNC", font=("Arial", 13, "bold")
        ).pack(side="right", padx=12)
        ctk.CTkLabel(

            self.titulo_bot_controls, text="TP+", font=("Arial", 13, "bold")
        ).pack(side="right", padx=15)
        ctk.CTkLabel(

            self.titulo_bot_controls, text="SL", font=("Arial", 13, "bold")
        ).pack(side="right", padx=18)

        self.bots_frame = ctk.CTkScrollableFrame(self.main_frame, height=400)
        self.bots_frame.pack(pady=5, padx=10, fill="both", expand=True)

        self.dashboard_frame = ctk.CTkFrame(self.main_frame)
        self.dashboard_frame.pack(pady=5, padx=10, fill="x")

        self.total_pnl_label = ctk.CTkLabel(
            self.dashboard_frame, text="Total PnL: 0 USDT", font=("Arial", 14, "bold")
        )
        self.total_pnl_label.pack(side="left", padx=10)

        self.total_trades_label = ctk.CTkLabel(
            self.dashboard_frame, text="Total Trades: 0", font=("Arial", 14, "bold")
        )
        self.total_trades_label.pack(side="left", padx=10)

        self.active_bots_label = ctk.CTkLabel(
            self.dashboard_frame, text="Bots Activos: 0", font=("Arial", 14, "bold")
        )
        self.active_bots_label.pack(side="left", padx=10)
        self.sync_status_label = ctk.CTkLabel(
            self.dashboard_frame,
            text="Sincronización: ⏺️ Última: --:--:--",
            font=("Arial", 14),
        )
        self.sync_status_label.pack(side="left", padx=10)


        self.exchange_var.set(list(EXCHANGES.keys())[0])
        self.load_exchange_api(self.exchange_var.get())

        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.after(2000, lambda: load_active_bots(self))

    def update_dashboard(self):
        total_pnl = sum([t.get("pnl", 0) for t in trading_history])
        bots_activos = len(self.bots)

        if total_pnl >= 0:
            pnl_text = f"+{round(total_pnl, 4)} USDT"
            pnl_color = "green"
        else:
            pnl_text = f"{round(total_pnl, 4)} USDT"
            pnl_color = "red"

        self.total_pnl_label.configure(
            text=f"Total PnL: {pnl_text}", text_color=pnl_color
        )
        self.total_trades_label.configure(
            text=f"Total Trades: {self.session_trades}")
        self.active_bots_label.configure(text=f"Bots Activos: {bots_activos}")

    def load_exchange_api(self, selected_exchange):
        exch_id = EXCHANGES[selected_exchange]
        exchange_data = self.config.get("exchanges", {}).get(exch_id, {})
        self.api_key_entry.delete(0, "end")
        self.api_secret_entry.delete(0, "end")
        self.api_key_entry.insert(0, exchange_data.get("apiKey", ""))
        self.api_secret_entry.insert(0, exchange_data.get("apiSecret", ""))

        connected = exchange_data.get("connected", False)
        if connected:
            threading.Thread(
                target=self.reconnect_exchange, args=(exch_id,), daemon=True
            ).start()
        self.status_circle.configure(
            text="🟢" if connected else "⚪", text_color="green" if connected else "gray"
        )
        self.connect_button.configure(
            text="Desconectar" if connected else "Conectar")

    def reconnect_exchange(self, exch_id):
        try:
            exchange_class = getattr(ccxt, exch_id)
            exchange = exchange_class(
                {
                    "apiKey": self.api_key_entry.get(),
                    "secret": self.api_secret_entry.get(),
                    "enableRateLimit": True,
                }
            )
            exchange.load_markets()
            self.exchange_instances[exch_id] = exchange
            load_active_bots(self)

        except Exception:
            traceback.print_exc()

    def toggle_connection(self):
        selected_exchange = self.exchange_var.get()
        exch_id = EXCHANGES[selected_exchange]

        if exch_id not in self.config.get("exchanges", {}):
            self.config.setdefault("exchanges", {})[exch_id] = {}

        if exch_id in self.exchange_instances:
            del self.exchange_instances[exch_id]
            self.config["exchanges"][exch_id]["connected"] = False
            save_config(self.config)
            self.status_circle.configure(text="⚪", text_color="gray")
            self.connect_button.configure(text="Conectar")
            return

        threading.Thread(
            target=self.connect_exchange, args=(exch_id,), daemon=True
        ).start()

    def connect_exchange(self, exch_id):
        try:
            exchange_class = getattr(ccxt, exch_id)
            exchange = exchange_class(
                {
                    "apiKey": self.api_key_entry.get(),
                    "secret": self.api_secret_entry.get(),
                    "enableRateLimit": True,
                }
            )
            exchange.load_markets()

            self.exchange_instances[exch_id] = exchange
            self.config["exchanges"][exch_id] = {
                "apiKey": self.api_key_entry.get(),
                "apiSecret": self.api_secret_entry.get(),
                "connected": True,
            }
            save_config(self.config)
            self.status_circle.configure(text="🟢", text_color="green")
            self.connect_button.configure(text="Desconectar")
        except Exception:
            traceback.print_exc()
            self.status_circle.configure(text="🔴", text_color="red")

    def start_bot(self):
        selected_exchange = self.exchange_var.get()
        exch_id = EXCHANGES[selected_exchange]
        exchange = self.exchange_instances.get(exch_id)

        if not exchange:
            return

        symbol = self.pair_var.get()
        try:
            exchange.load_markets()
            if symbol not in exchange.symbols:
                print(f"❌ El par {symbol} no existe en este exchange.")
                return

            monto = float(self.monto_entry.get() or 0)
            tp_pct = float(self.tp_entry.get() or 0)
            sep_pct = float(self.sep_entry.get() or 0)
            os_num = int(float(self.os_entry.get() or 0))

            bot = BotFrame(
                self.bots_frame,
                symbol,
                monto,
                tp_pct,
                sep_pct,
                os_num,
                exchange,
                self.remove_bot,
            )

            bot.pack(pady=2, fill="x")
            self.bots.append(bot)
            self.update_dashboard()

            # ✅ También iniciar el bot en el servidor
            try:
                bot_data = {
                    "id": f"{symbol}-{exch_id}",
                    "symbol": symbol,
                    "monto": monto,
                    "tp_pct": tp_pct,
                    "sep_pct": sep_pct,
                    "os_num": os_num,
                    "exchange": exch_id,
                    "vnc_total": 0,
                    "vnc_total_cost": 0,
                    "total_trades": 0,
                    "contador_ciclos": 0,
                    "dca_pct": "0.5%",
                    "sl": "",
                    "tp_plus": "",
                    "reiniciar": True,
                    "reiniciar_os": True,
                    "apiKey": self.api_key_entry.get(),
                    "apiSecret": self.api_secret_entry.get(),
  
                }
                requests.post(
                    "https://bot-server-cg3g.onrender.com/api/start_bot", json=bot_data)
                print(f"🚀 Bot enviado al servidor: {bot_data['id']}")
            except Exception as e:
                print(f"[SERVER START ERROR] {e}")

        except Exception as e:
            print(f"Error iniciando bot para {symbol}: {e}")

    def remove_bot(self, bot):
        if bot in self.bots:
            self.bots.remove(bot)
        self.update_dashboard()

    def on_closing(self):
        save_config(self.config)
        save_active_bots(self)
        self.destroy()

    def abrir_historial(self):
        self.main_frame.pack_forget()
        self.historial_frame = ctk.CTkFrame(self)
        self.historial_frame.pack(fill="both", expand=True)
        self.historial_frame.configure(fg_color="white")

        # Top bar con botón volver y exportar
        top_bar = ctk.CTkFrame(self.historial_frame, fg_color="white")
        top_bar.pack(fill="x", pady=10, padx=20)

        volver_btn = ctk.CTkButton(
            top_bar, text="← Volver", command=self.cerrar_historial
        )
        volver_btn.pack(side="left")

        export_btn = ctk.CTkButton(
            top_bar, text="📤 Exportar Excel", command=save_trading_history
        )
        export_btn.pack(side="right", padx=5)

        # Tabla de historial principal
        self.historial_table = ctk.CTkScrollableFrame(
            self.historial_frame, height=500, fg_color="white"
        )
        self.historial_table.pack(
            padx=20, pady=5, fill="both", expand=True, side="left"
        )

        self.paginador_frame = ctk.CTkFrame(
            self.historial_frame, fg_color="white")
        self.paginador_frame.pack(pady=5)

        self.historial_pagina_actual = 0
        self.historial_por_pagina = 50
        self.cargar_historial_pagina()

        # Panel lateral de PnL por par
        self.panel_pnl_resumen = ctk.CTkFrame(
            self.historial_frame, width=240, fg_color="white"
        )
        self.panel_pnl_resumen.pack(
            side="right", fill="y", padx=(0, 20), pady=10)
        self.actualizar_panel_pnl()

    def cerrar_historial(self):
        self.historial_frame.pack_forget()
        self.main_frame.pack(fill="both", expand=True)

    def cargar_historial_pagina(self):
        for widget in self.historial_table.winfo_children():
            widget.destroy()

        inicio = self.historial_pagina_actual * self.historial_por_pagina
        fin = inicio + self.historial_por_pagina
        pagina_trades = trading_history[::-1][inicio:fin]

        for idx, t in enumerate(pagina_trades, start=1):
            linea = (
                f"{idx+inicio}. {t['symbol']} | {t['datetime']} | Ent: {t['entry_price']} | Sal: {t['exit_price']} | "
                f"MO: {t['amount']} | FBuy: {t['fee_buy']} | FSell: {t['fee_sell']} | PnL: {t['pnl']}"
            )
            color = "green" if t["pnl"] >= 0 else "red"
            lbl = ctk.CTkLabel(
                self.historial_table,
                text=linea,
                text_color=color,
                font=("Arial", 12),
                anchor="w",
            )
            lbl.pack(fill="x", padx=10)

            divider = ctk.CTkLabel(
                self.historial_table, text="─" * 300, text_color="#cccccc"
            )
            divider.pack(fill="x", padx=0, pady=(0, 4))

        # Botones de paginación
        for widget in self.paginador_frame.winfo_children():
            widget.destroy()

        anterior = ctk.CTkButton(
            self.paginador_frame, text="← Anterior", command=self.pagina_anterior
        )
        siguiente = ctk.CTkButton(
            self.paginador_frame, text="Siguiente →", command=self.pagina_siguiente
        )
        anterior.pack(side="left", padx=10)
        siguiente.pack(side="left", padx=10)

    def pagina_anterior(self):
        if self.historial_pagina_actual > 0:
            self.historial_pagina_actual -= 1
            self.cargar_historial_pagina()

    def pagina_siguiente(self):
        max_pagina = len(trading_history) // self.historial_por_pagina
        if self.historial_pagina_actual < max_pagina:
            self.historial_pagina_actual += 1
            self.cargar_historial_pagina()

    def actualizar_panel_pnl(self):
        for widget in self.panel_pnl_resumen.winfo_children():
            widget.destroy()

        filtro_frame = ctk.CTkFrame(self.panel_pnl_resumen, fg_color="white")
        filtro_frame.pack(pady=(5, 10))

        self.pnl_periodo = tk.StringVar(value="Día")

        for periodo in ["Día", "Semana", "Mes", "Año"]:
            btn = ctk.CTkRadioButton(
                filtro_frame,
                text=periodo,
                variable=self.pnl_periodo,
                value=periodo,
                command=self.cargar_resumen_pnl,
                fg_color="white",
            )
            btn.pack(side="left", padx=3)

        self.lista_resumen = ctk.CTkScrollableFrame(
            self.panel_pnl_resumen, fg_color="white", width=240, height=500
        )
        self.lista_resumen.pack()

        self.cargar_resumen_pnl()

    def cargar_resumen_pnl(self):
        for widget in self.lista_resumen.winfo_children():
            widget.destroy()

        ahora = datetime.now()
        periodo = self.pnl_periodo.get()

        def filtro_func(t):
            return True

        if periodo == "Día":

            def filtro_func(t):
                return (
                    datetime.strptime(
                        t["datetime"], "%d-%m-%Y %H:%M:%S").date()
                    == ahora.date()
                )

        elif periodo == "Semana":
            semana_actual = ahora.isocalendar()[1]

            def filtro_func(t):
                return (
                    datetime.strptime(t["datetime"], "%d-%m-%Y %H:%M:%S").isocalendar()[
                        1
                    ]
                    == semana_actual
                )

        elif periodo == "Mes":

            def filtro_func(t):
                return (
                    datetime.strptime(t["datetime"], "%d-%m-%Y %H:%M:%S").month
                    == ahora.month
                )

        elif periodo == "Año":

            def filtro_func(t):
                return (
                    datetime.strptime(t["datetime"], "%d-%m-%Y %H:%M:%S").year
                    == ahora.year
                )

        trades_filtrados = [t for t in trading_history if filtro_func(t)]
        resumen = {}

        for t in trades_filtrados:
            par = t["symbol"]
            resumen[par] = resumen.get(par, 0) + t["pnl"]

        resumen_ordenado = sorted(resumen.items(), key=lambda x: -x[1])

        for i, (par, pnl) in enumerate(resumen_ordenado, start=1):
            color = "green" if pnl >= 0 else "red"
            texto = f"#{i}. {par}   {round(pnl, 6)}"
            lbl = ctk.CTkLabel(
                self.lista_resumen,
                text=texto,
                text_color=color,
                anchor="w",
                font=("Arial", 12),
            )
            lbl.pack(anchor="w", padx=10, pady=1)


def save_active_bots(app):
    bots_data = []
    for bot in app.bots:
        bots_data.append({
            "symbol": bot.symbol,
            "monto": bot.monto,
            "tp_pct": bot.tp_pct,
            "sep_pct": bot.sep_pct,
            "os_num": bot.os_num,
            "exchange": bot.exchange.id,
            "vnc_total": bot.vnc_total,
            "vnc_total_cost": bot.vnc_total_cost,
            "total_trades": bot.total_trades,
            "contador_ciclos": bot.contador_ciclos,
            "dca_pct": bot.dca_sell_pct_var.get(),
            "sl": bot.sl_pct_var.get(),
            "tp_plus": bot.tp_plus_var.get(),
            "reiniciar": bot.reiniciar_var.get(),
            "reiniciar_os": bot.reiniciar_os_var.get(),
        })
    with open("data/active_bots.json", "w") as f:
        json.dump(bots_data, f, indent=2)


def load_active_bots(app):
    if not os.path.exists("data/active_bots.json"):
        return
    try:
        with open("data/active_bots.json", "r") as f:
            bots_data = json.load(f)
        for data in bots_data:
            exch_id = data["exchange"]
            if exch_id not in app.exchange_instances:
                continue
            exchange = app.exchange_instances[exch_id]
            bot = BotFrame(
                app.bots_frame,
                data["symbol"],
                data["monto"],
                data["tp_pct"],
                data["sep_pct"],
                data["os_num"],
                exchange,
                app.remove_bot,
            )
            bot.vnc_total = data["vnc_total"]
            bot.vnc_total_cost = data["vnc_total_cost"]
            bot.total_trades = data["total_trades"]
            bot.contador_ciclos = data["contador_ciclos"]
            bot.dca_sell_pct_var.set(data["dca_pct"])
            bot.sl_pct_var.set(data["sl"])
            bot.tp_plus_var.set(data["tp_plus"])
            bot.reiniciar_var.set(data["reiniciar"])
            bot.reiniciar_os_var.set(data["reiniciar_os"])
            if data["reiniciar"]:
                bot.reiniciar_checkbox.select()
            if data["reiniciar_os"]:
                bot.reiniciar_os_checkbox.select()
            bot.pack(pady=2, fill="x")
            bot.update_labels()
            app.bots.append(bot)
    except Exception as e:
        print(f"[ERROR] Cargando bots activos: {e}")





if __name__ == "__main__":
    app = BotDCAApp()

    # ---- CLIENTE DE SINCRONIZACIÓN REMOTA ----
    import requests
    import threading
    import time
    from datetime import datetime

    class ServerSync:
        def __init__(self, app, url_base="https://bot-server-cg3g.onrender.com"):
            self.app = app
            self.url = url_base
            threading.Thread(target=self.start_sync, daemon=True).start()

        def start_sync(self):
            try:
                self.send_config()
                while True:
                    time.sleep(10)
                    self.sync_bots()
            except Exception as e:
                print(f"[ServerSync ERROR] {e}")

        def sync_bots(self):
            try:
                for bot in self.app.bots:
                    bot_data = {
                        "id": f"{bot.symbol}-{bot.exchange_id}",
                        "tp_pct": bot.tp_pct,
                        "tp_plus": bot.tp_plus_var.get(),
                        "sl": bot.sl_pct_var.get(),
                        "dca_pct": bot.dca_sell_pct_var.get(),
                        "reiniciar": bot.reiniciar_var.get(),
                        "reiniciar_os": bot.reiniciar_os_var.get(),
                        "monto": bot.monto,
                        "sep_pct": bot.sep_pct,
                        "os_num": bot.os_num
                    }
                    requests.post(f"{self.url}/api/update_bot", json=bot_data)

                hora = datetime.now().strftime("%H:%M:%S")
                print(f"🔁 Bots sincronizados con el servidor ({hora})")
                self.app.sync_status_label.configure(
                    text=f"Sincronización: 🟢 Última: {hora}",
                    text_color="green"
                )

            except Exception as e:
                print(f"[SYNC ERROR] {e}")
                hora = datetime.now().strftime("%H:%M:%S")
                self.app.sync_status_label.configure(
                    text=f"Sincronización: 🔴 Última: {hora}",
                    text_color="red"
                )

        def send_config(self):
            try:
                requests.post(f"{self.url}/api/config", json=self.app.config)
                print("✔️ Enviado /api/config")
            except Exception as e:
                print(f"[CONFIG ERROR] {e}")

    ServerSync(app)
    app.mainloop()
